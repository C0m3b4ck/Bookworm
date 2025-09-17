import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

FILE_NAME_EN = "books_en.xlsx"
FILE_NAME_PL = "books_pl.xlsx"

def select_language():
    language = str(input("Select language / wybierz język [EN/PL]: "))
    if language.lower() in ["en", "e"]:
        print("English selected.\n")
        welcome_en()
    elif language.lower() in ["pl", "p"]:
        print("Wybrano polski.\n")
        welcome_pl()
    else:
        print("Unknown language / nieznany język")
        select_language()

def welcome_en():
    os.system('cls' if os.name == 'nt' else 'clear')
    print("-/Welcome to Bookworm!\\-")
    print("Version Beta 1.1")
    print("Made by C0m3b4ck under APL 2.0 license\n")
    menu_en()

def welcome_pl():
    os.system('cls' if os.name == 'nt' else 'clear')
    print("-/Witamy w Bookworm!\\-")
    print("Wersja Beta 1.1")
    print("Stworzone przez C0m3b4ck pod licencją APL 2.0\n")
    menu_pl()

def load_or_create_workbook(lang="EN"):
    file_name = FILE_NAME_EN if lang == "EN" else FILE_NAME_PL
    if not os.path.exists(file_name):
        if lang == "EN":
            create = input(f"Excel file '{file_name}' not found. Create new? [Y/N]: ").strip().lower()
        else:
            create = input(f"Nie znaleziono pliku excela '{file_name}'. Utworzyć nowy? [T/N]: ").strip().lower()
        if create == 'y' or create == 't':
            wb = Workbook()
            ws = wb.active
            ws.title = "Books"
            headers = ["ID", "Title", "Author", "Year", "Genre"]

            header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            ws.append(headers)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                ws.column_dimensions[get_column_letter(col_num)].width = 15
            wb.save(file_name)
            print(f"New Excel file '{file_name}' created.\n")
            return wb, ws, True
        else:
            if lang == "EN":
                print("Excel file needed to proceed. Exiting.")
            else:
                print("Plik excela wymagany do kontynuacji. Kończę działanie.")
            sys.exit()
    wb = load_workbook(file_name)
    if "Books" not in wb.sheetnames:
        ws = wb.create_sheet("Books")
        headers = ["ID", "Title", "Author", "Year", "Genre"]
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        wb.save(file_name)
    else:
        ws = wb["Books"]
    return wb, ws, False

def get_existing_ids(ws):
    return [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0] is not None]

def find_books_by_name(ws, name):
    books = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] and name.lower() in str(row[1]).lower():
            books.append({
                "ID": row[0],
                "Title": row[1],
                "Author": row[2],
                "Year": row[3],
                "Genre": row[4]
            })
    return books

def find_book_by_id(ws, id_):
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        cell_id = row[0].value
        if cell_id == id_:
            return idx, row
    return None, None

def display_book(book):
    print(f"ID: {book['ID']}")
    print(f"Tytuł: {book['Title']}")
    print(f"Autor: {book['Author']}")
    print(f"Rok: {book['Year']}")
    print(f"Gatunek: {book['Genre']}")
    print("-" * 20)

# Polish version book input allowing ID 0 duplicates
def wczytaj_dane_ksiazki(existing=None, ws=None):
    existing_ids = get_existing_ids(ws) if ws else []
    entered_id = None

    while True:
        try:
            id_input = input(f"ID [{existing['ID'] if existing else 'wpisz unikalny numer >= 0 (0 może się powtarzać)'}]: ").strip()
            if id_input == "" and existing:
                entered_id = existing['ID']
                break
            entered_id = int(id_input)
            if entered_id < 0:
                print("ID musi być większe lub równe 0.")
                continue
            # allow duplicates only if ID = 0
            if entered_id != 0 and entered_id in existing_ids and (not existing or entered_id != existing['ID']):
                print("ID już istnieje. Proszę podać unikalny ID lub 0, który może się powtarzać.")
                continue
            break
        except ValueError:
            print("ID musi być liczbą.")

    title = input(f"Tytuł [{existing['Title'] if existing else ''}]: ").strip()
    if not title and existing:
        title = existing['Title']
    author = input(f"Autor [{existing['Author'] if existing else ''}]: ").strip()
    if not author and existing:
        author = existing['Author']
    year = input(f"Rok [{existing['Year'] if existing else ''}]: ").strip()
    if not year and existing:
        year = existing['Year']
    else:
        if year and not year.isdigit():
            print("Rok musi być liczbą, pozostawiam puste.")
            year = ""
    genre = input(f"Gatunek [{existing['Genre'] if existing else ''}]: ").strip()
    if not genre and existing:
        genre = existing['Genre']

    return {
        "ID": entered_id,
        "Title": title,
        "Author": author,
        "Year": year,
        "Genre": genre
    }

# English book input allowing ID 0 duplicates
def prompt_book_details(existing=None, require_id=False, ws=None):
    entered_id = None
    if require_id:
        existing_ids = get_existing_ids(ws) if ws else []
        while True:
            try:
                id_input = input(f"ID [{existing['ID'] if existing else 'must enter a unique number >=0 (0 may repeat)'}]: ").strip()
                if id_input == "" and existing:
                    entered_id = existing['ID']
                    break
                entered_id = int(id_input)
                if entered_id < 0:
                    print("ID must be 0 or higher.")
                    continue
                if entered_id != 0 and entered_id in existing_ids and (not existing or entered_id != existing['ID']):
                    print("ID already exists. Enter a unique ID or 0 which may repeat.")
                    continue
                break
            except ValueError:
                print("ID must be a number.")
    else:
        entered_id = existing['ID'] if existing else None

    title = input(f"Title [{existing['Title'] if existing else ''}]: ").strip()
    if not title and existing:
        title = existing['Title']
    author = input(f"Author [{existing['Author'] if existing else ''}]: ").strip()
    if not author and existing:
        author = existing['Author']
    year = input(f"Year [{existing['Year'] if existing else ''}]: ").strip()
    if not year and existing:
        year = existing['Year']
    else:
        if year and not year.isdigit():
            print("Year must be a number, storing as empty.")
            year = ""
    genre = input(f"Genre [{existing['Genre'] if existing else ''}]: ").strip()
    if not genre and existing:
        genre = existing['Genre']

    return {
        "ID": entered_id,
        "Title": title,
        "Author": author,
        "Year": year,
        "Genre": genre
    }

# Polish menu and functions using modified input function
def menu_pl():
    os.system('cls' if os.name == 'nt' else 'clear')
    print("Wybierz opcje:")
    print("1. Dodaj książkę")
    print("2. Obejrzyj/modyfikuj książki")
    print("3. Zobacz Pomoc")
    print("4. Opusć program")
    menu_choice = str(input("Wybierz opcję: "))

    wb, ws, just_created = load_or_create_workbook(lang="PL")
    if just_created:
        dodaj_ksiazke(wb, ws)
        return
    if menu_choice == "1":
        dodaj_ksiazke(wb, ws)
    elif menu_choice == "2":
        przeglad_modyfikacja(wb, ws)
    elif menu_choice == "3":
        os.system('cls' if os.name == 'nt' else 'clear')
        print("Pomoc")
        print("Opcja '1' - dodanie książki do księgozbioru")
        print("Opcja '2' - obejrzenie/modyfikowanie księgozbioru, obejmuje usuwanie lub zmienianie danych")
        print("Opcja '3' - wyświetla informacje o programie")
        print("Opcja '4' - opuszcza program")
        input("Wpisz losowy znak aby wrócić do menu opcji: ")
        menu_pl()
    elif menu_choice == "4":
        os.system('cls' if os.name == 'nt' else 'clear')
        print("Opuszczanie programu...")
        sys.exit()
    else:
        print("Nieprawidłowa opcja.")
        input("Wciśnij Enter aby wrócić do menu...")
        menu_pl()

def dodaj_ksiazke(wb, ws):
    os.system('cls' if os.name == 'nt' else 'clear')
    print("Dodawanie książki")
    new_book = wczytaj_dane_ksiazki(ws=ws)
    print("\nWpisano:")
    display_book(new_book)
    confirm = input("Potwierdź dodanie książki? [T/N]: ").strip().lower()
    if confirm == 't':
        ws.append([new_book["ID"], new_book["Title"], new_book["Author"], new_book["Year"], new_book["Genre"]])
        wb.save(FILE_NAME_PL)
        print("Książka dodana pomyślnie.")
    else:
        print("Dodanie anulowane.")
    input("Wciśnij Enter aby wrócić do menu...")
    menu_pl()

def przeglad_modyfikacja(wb, ws):
    os.system('cls' if os.name == 'nt' else 'clear')
    print("Przeglądanie/modyfikowanie książek")
    print("Przeglądaj po:")
    print("1. ID")
    print("2. Nazwa")
    choice = input("Wybierz opcję przeglądania: ")
    if choice == '1':
        try:
            book_id = int(input("Wpisz ID książki: "))
        except ValueError:
            print("Nieprawidłowe ID, musi być liczba.")
            input("Wciśnij Enter aby wrócić do menu...")
            menu_pl()
            return
        idx, row = find_book_by_id(ws, book_id)
        if row is None:
            print(f"Nie znaleziono książki z ID {book_id}.")
            input("Wciśnij Enter aby wrócić do menu...")
            menu_pl()
            return
        book = { "ID": row[0].value, "Title": row[1].value, "Author": row[2].value, "Year": row[3].value, "Genre": row[4].value }
        print("Znaleziono książkę:")
        display_book(book)
    elif choice == '2':
        name = input("Wpisz frazę nazwy książki do wyszukania: ").strip()
        results = find_books_by_name(ws, name)
        if not results:
            print(f"Nie znaleziono książek pasujących do '{name}'.")
            input("Wciśnij Enter aby wrócić do menu...")
            menu_pl()
            return
        print(f"Znaleziono {len(results)} książkę/ki:")
        for b in results:
            display_book(b)
        try:
            select_id = int(input("Wpisz ID książki do modyfikacji lub 0 aby anulować: "))
        except ValueError:
            print("Nieprawidłowy wpis.")
            input("Wciśnij Enter aby wrócić do menu...")
            menu_pl()
            return
        if select_id == 0:
            menu_pl()
            return
        idx, row = find_book_by_id(ws, select_id)
        if row is None:
            print(f"Nie znaleziono książki z ID {select_id}.")
            input("Wciśnij Enter aby wrócić do menu...")
            menu_pl()
            return
        book = { "ID": row[0].value, "Title": row[1].value, "Author": row[2].value, "Year": row[3].value, "Genre": row[4].value }
        print("Wybrano książkę:")
        display_book(book)
    else:
        print("Nieprawidłowy wybór.")
        input("Wciśnij Enter aby wrócić do menu...")
        menu_pl()
        return

    print("Opcje:")
    print("1. Modyfikuj książkę")
    print("2. Usuń książkę")
    print("3. Wróć do menu")
    action = input("Wybierz akcję: ")
    if action == "1":
        modified_data = wczytaj_dane_ksiazki(existing=book, ws=ws)
        print("\nZmodyfikowane dane:")
        display_book(modified_data)
        confirm_mod = input("Potwierdź modyfikację? [T/N]: ").strip().lower()
        if confirm_mod == 't':
            ws.cell(row=idx, column=1).value = modified_data["ID"]
            ws.cell(row=idx, column=2).value = modified_data["Title"]
            ws.cell(row=idx, column=3).value = modified_data["Author"]
            ws.cell(row=idx, column=4).value = modified_data["Year"]
            ws.cell(row=idx, column=5).value = modified_data["Genre"]
            wb.save(FILE_NAME_PL)
            print("Książka zmodyfikowana pomyślnie.")
        else:
            print("Modyfikacja anulowana.")
        input("Wciśnij Enter aby wrócić do menu...")
        menu_pl()
    elif action == "2":
        confirm_del = input(f"Czy na pewno chcesz usunąć książkę ID {book['ID']}? [T/N]: ").strip().lower()
        if confirm_del == 't':
            ws.delete_rows(idx)
            wb.save(FILE_NAME_PL)
            print("Książka usunięta.")
        else:
            print("Usunięcie anulowane.")
        input("Wciśnij Enter aby wrócić do menu...")
        menu_pl()
    else:
        menu_pl()

# English menus and functions with same ID 0 allowance

def menu_en():
    os.system('cls' if os.name == 'nt' else 'clear')
    print("Select options: ")
    print("1. Add new book")
    print("2. See/modify books")
    print("3. See Help")
    print("4. Exit program")
    menu_choice = str(input("Select option: "))

    wb, ws, just_created = load_or_create_workbook(lang="EN")
    if just_created:
        add_new_book_en(wb, ws)
        return
    if menu_choice == "1":
        add_new_book_en(wb, ws)
    elif menu_choice == "2":
        see_modify_books_en(wb, ws)
    elif menu_choice == "3":
        os.system('cls' if os.name == 'nt' else 'clear')
        print("Help")
        print("Option '1' - adding book to book list")
        print("Option '2' - seeing/modifying books, covers removal or modification of data")
        print("Option '3' - shows information about the program")
        print("Option '4' - leaves the program")
        input("Input random character to view menu: ")
        menu_en()
    elif menu_choice == "4":
        os.system('cls' if os.name == 'nt' else 'clear')
        print("Exiting program...")
        sys.exit()
    else:
        print("Invalid option.")
        input("Press Enter to return to menu...")
        menu_en()

def add_new_book_en(wb, ws):
    os.system('cls' if os.name == 'nt' else 'clear')
    print("Add new book")
    new_book = prompt_book_details(require_id=True, ws=ws)
    print("\nYou entered:")
    display_book(new_book)
    confirm = input("Confirm adding book? [Y/N]: ").strip().lower()
    if confirm == 'y':
        ws.append([new_book["ID"], new_book["Title"], new_book["Author"], new_book["Year"], new_book["Genre"]])
        wb.save(FILE_NAME_EN)
        print("Book added successfully.")
    else:
        print("Addition cancelled.")
    input("Press Enter to return to menu...")
    menu_en()

def see_modify_books_en(wb, ws):
    os.system('cls' if os.name == 'nt' else 'clear')
    print("See/modify books")
    print("Browse by:")
    print("1. ID")
    print("2. Name")
    choice = input("Choose browsing option: ")
    if choice == '1':
        try:
            book_id = int(input("Enter Book ID: "))
        except ValueError:
            print("Invalid ID, must be a number.")
            input("Press Enter to return to menu...")
            menu_en()
            return
        idx, row = find_book_by_id(ws, book_id)
        if row is None:
            print(f"No book with ID {book_id} found.")
            input("Press Enter to return to menu...")
            menu_en()
            return
        book = { "ID": row[0].value, "Title": row[1].value, "Author": row[2].value, "Year": row[3].value, "Genre": row[4].value }
        print("Book found:")
        display_book(book)
    elif choice == '2':
        name = input("Enter part of the book name to search: ").strip()
        results = find_books_by_name(ws, name)
        if not results:
            print(f"No books matching '{name}' found.")
            input("Press Enter to return to menu...")
            menu_en()
            return
        print(f"Found {len(results)} book(s):")
        for b in results:
            display_book(b)
        try:
            select_id = int(input("Enter ID of the book to modify, or 0 to cancel: "))
        except ValueError:
            print("Invalid input.")
            input("Press Enter to return to menu...")
            menu_en()
            return
        if select_id == 0:
            menu_en()
            return
        idx, row = find_book_by_id(ws, select_id)
        if row is None:
            print(f"No book with ID {select_id} found.")
            input("Press Enter to return to menu...")
            menu_en()
            return
        book = { "ID": row[0].value, "Title": row[1].value, "Author": row[2].value, "Year": row[3].value, "Genre": row[4].value }
        print("Selected book:")
        display_book(book)
    else:
        print("Invalid choice.")
        input("Press Enter to return to menu...")
        menu_en()
        return

    print("Options:")
    print("1. Modify book")
    print("2. Delete book")
    print("3. Return to menu")
    action = input("Select action: ")
    if action == "1":
        modified_data = prompt_book_details(existing=book, require_id=True, ws=ws)
        print("\nModified details:")
        display_book(modified_data)
        confirm_mod = input("Confirm modification? [Y/N]: ").strip().lower()
        if confirm_mod == 'y':
            ws.cell(row=idx, column=1).value = modified_data["ID"]
            ws.cell(row=idx, column=2).value = modified_data["Title"]
            ws.cell(row=idx, column=3).value = modified_data["Author"]
            ws.cell(row=idx, column=4).value = modified_data["Year"]
            ws.cell(row=idx, column=5).value = modified_data["Genre"]
            wb.save(FILE_NAME_EN)
            print("Book modified successfully.")
        else:
            print("Modification cancelled.")
        input("Press Enter to return to menu...")
        menu_en()
    elif action == "2":
        confirm_del = input(f"Are you sure you want to delete book ID {book['ID']}? [Y/N]: ").strip().lower()
        if confirm_del == 'y':
            ws.delete_rows(idx)
            wb.save(FILE_NAME_EN)
            print("Book deleted.")
        else:
            print("Deletion cancelled.")
        input("Press Enter to return to menu...")
        menu_en()
    else:
        menu_en()

if __name__ == "__main__":
    select_language()
