import os
import json
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import re
import subprocess


FILE_NAME_EN = "books_en.xlsx"
FILE_NAME_PL = "books_pl.xlsx"
REMOVED_FILE_EN = "removed_en.xlsx"
REMOVED_FILE_PL = "removed_pl.xlsx"
SETTINGS_FILE = "settings.json"
THEMES_FOLDER = "themes"
DEFAULT_SETTINGS = {
    "default_language": None,  # if None, prompt each time
    "theme": "classic_blue"
}
DEFAULT_THEMES = {
    "classic_blue": {
        "BTN_BG": "#add8e6",         # light blue
        "BTN_FG": "#333333",         # dark grey
        "BTN_HOVER_BG": "#6495ed",  # cornflower blue
        "BTN_HOVER_FG": "#f0f0f0",  # lighter text
    },
    "dark_mode": {
        "BTN_BG": "#444444",
        "BTN_FG": "#ffffff",
        "BTN_HOVER_BG": "#666666",
        "BTN_HOVER_FG": "#ffff00",
    }
}


def find_book_by_id(ws, id_):
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        cell_id = row[0].value
        if cell_id == id_:
            return idx, row
    return None, None


def find_latest_version_executable(prefix):
    pattern = re.compile(rf"^{prefix}_v(\d+)\.exe$", re.IGNORECASE)
    best_version = -1
    best_path = None
    for root, dirs, files in os.walk("."):
        for file in files:
            match = pattern.match(file)
            if match:
                ver = int(match.group(1))
                if ver > best_version:
                    best_version = ver
                    best_path = os.path.join(root, file)
    return best_path, best_version


class BookwormApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Bookworm")
        self.geometry("750x520")
        self.minsize(600, 450)
        self.settings = self.load_settings()
        self.themes = DEFAULT_THEMES.copy()
        self.load_custom_themes()
        theme_name = self.settings.get("theme", "classic_blue")
        self.set_theme(theme_name)
        self.lang = self.settings.get("default_language")
        self.file_name = None
        self.removed_file_name = None
        self.wb = None
        self.ws = None
        if self.lang not in ("EN", "PL"):
            self.create_language_selection()
        else:
            self.start_app(self.lang)

    def load_settings(self):
        if not os.path.exists(SETTINGS_FILE):
            self.save_settings(DEFAULT_SETTINGS)
            return DEFAULT_SETTINGS.copy()
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return DEFAULT_SETTINGS.copy()

    def save_settings(self, settings):
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(settings, f, indent=4)

    def load_custom_themes(self):
        if not os.path.exists(THEMES_FOLDER):
            os.makedirs(THEMES_FOLDER)
            return
        for file in os.listdir(THEMES_FOLDER):
            if file.endswith(".json"):
                try:
                    path = os.path.join(THEMES_FOLDER, file)
                    with open(path, "r", encoding="utf-8") as f:
                        theme_data = json.load(f)
                    keys = {"BTN_BG", "BTN_FG", "BTN_HOVER_BG", "BTN_HOVER_FG"}
                    if keys.issubset(theme_data.keys()):
                        name = os.path.splitext(file)[0]
                        self.themes[name] = theme_data
                except Exception:
                    pass

    def set_theme(self, name):
        theme = self.themes.get(name, self.themes["classic_blue"])
        self.BTN_BG = theme["BTN_BG"]
        self.BTN_FG = theme["BTN_FG"]
        self.BTN_HOVER_BG = theme["BTN_HOVER_BG"]
        self.BTN_HOVER_FG = theme["BTN_HOVER_FG"]

    def create_high_contrast_button(self, parent, text, command=None):
        btn = tk.Button(parent, text=text, bg=self.BTN_BG, fg=self.BTN_FG,
                        activebackground=self.BTN_HOVER_BG, activeforeground=self.BTN_HOVER_FG,
                        font=("Segoe UI", 12, "bold"), command=command, relief="raised", bd=3)
        btn.bind("<Enter>", lambda e: btn.config(bg=self.BTN_HOVER_BG, fg=self.BTN_HOVER_FG))
        btn.bind("<Leave>", lambda e: btn.config(bg=self.BTN_BG, fg=self.BTN_FG))
        return btn

    def create_language_selection(self):
        for widget in self.winfo_children():
            widget.destroy()
        tk.Label(self, text="Select language / wybierz język", font=("Arial", 18)).grid(row=0, column=0, columnspan=2, pady=20)
        btn_en = self.create_high_contrast_button(self, "English", lambda: self.set_language_and_start("EN"))
        btn_en.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
        btn_pl = self.create_high_contrast_button(self, "Polski", lambda: self.set_language_and_start("PL"))
        btn_pl.grid(row=1, column=1, sticky="nsew", padx=10, pady=5)
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)

    def set_language_and_start(self, lang):
        self.lang = lang
        self.settings["default_language"] = lang
        self.save_settings(self.settings)
        self.start_app(lang)

    def start_app(self, lang):
        self.lang = lang
        self.file_name = FILE_NAME_EN if lang == "EN" else FILE_NAME_PL
        self.removed_file_name = REMOVED_FILE_EN if lang == "EN" else REMOVED_FILE_PL
        self.load_or_create_workbook()
        self.create_main_menu()

    def load_or_create_workbook(self):
        if not os.path.exists(self.file_name):
            if messagebox.askyesno(
                "Create file" if self.lang == "EN" else "Utwórz plik",
                f"{'Excel file not found. Create it?' if self.lang == 'EN' else 'Plik excela nie znaleziony. Utworzyć nowy?'}"):
                self.wb = Workbook()
                self.ws = self.wb.active
                self.ws.title = "Books"
                headers = ["ID", "Title", "Author", "Year", "Genre", "Status"]
                header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                self.ws.append(headers)
                for col_num in range(1, len(headers) + 1):
                    cell = self.ws.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    self.ws.column_dimensions[get_column_letter(col_num)].width = 15
                self.wb.save(self.file_name)
            else:
                messagebox.showinfo("Info", "Exiting." if self.lang == "EN" else "Kończenie programu.")
                self.quit()
                return
        self.wb = load_workbook(self.file_name)
        if "Books" not in self.wb.sheetnames:
            self.ws = self.wb.create_sheet("Books")
            headers = ["ID", "Title", "Author", "Year", "Genre", "Status"]
            header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            self.ws.append(headers)
            for col_num in range(1, len(headers) + 1):
                cell = self.ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                self.ws.column_dimensions[get_column_letter(col_num)].width = 15
            self.wb.save(self.file_name)
        else:
            self.ws = self.wb["Books"]
        if not os.path.exists(self.removed_file_name):
            wb_removed = Workbook()
            ws_removed = wb_removed.active
            ws_removed.title = "RemovedBooks"
            headers = ["ID", "Title", "Author", "Year", "Genre"]
            ws_removed.append(headers)
            for col_num in range(1, len(headers) + 1):
                cell = ws_removed.cell(row=1, column=col_num)
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                ws_removed.column_dimensions[get_column_letter(col_num)].width = 15
            wb_removed.save(self.removed_file_name)

    def create_main_menu(self):
        for widget in self.winfo_children():
            widget.destroy()
        credits_text = ("Made by C0m3b4ck under APL 2.0 license"
                        if self.lang == "EN" else "Stworzone przez C0m3b4ck pod licencją APL 2.0")
        label_text = "Main Menu" if self.lang == "EN" else "Menu główne"
        tk.Label(self, text=label_text, font=("Arial", 20)).grid(row=0, column=0, columnspan=3, pady=10)
        tk.Label(self, text=credits_text, font=("Arial", 10, "italic")).grid(row=1, column=0, columnspan=3, pady=5)

        options = {
            "EN": ["Add New Book", "See/Modify Books", "Help", "Update", "Settings", "Exit"],
            "PL": ["Dodaj książkę", "Obejrzyj/modyfikuj książki", "Pomoc", "Aktualizuj", "Ustawienia", "Zakończ"]
        }
        commands = [
            lambda: self.add_new_book_form(),
            lambda: self.see_modify_books(),
            lambda: self.show_help(),
            self.run_updater,
            self.show_settings,
            self.quit,
        ]
        for i, (option, command) in enumerate(zip(options[self.lang], commands), start=2):
            btn = self.create_high_contrast_button(self, option, command=command)
            btn.grid(row=i, column=0, columnspan=3, sticky="nsew", padx=50, pady=7)
        for i in range(2, 8):
            self.grid_rowconfigure(i, weight=1)
        for j in range(3):
            self.grid_columnconfigure(j, weight=1)

    def run_updater(self):
        updater_file, ver = find_latest_version_executable("updater")
        if not updater_file:
            messagebox.showinfo(
                "Updater Not Found" if self.lang == "EN" else "Updater nie znaleziony",
                "Updater executable not found. Please install/update manually."
                if self.lang == "EN"
                else "Nie znaleziono programu aktualizującego. Proszę zainstalować/zaktualizować ręcznie."
            )
            return
        try:
            subprocess.Popen([os.path.abspath(updater_file)], shell=True)
        except Exception as e:
            messagebox.showerror(
                "Error" if self.lang == "EN" else "Błąd",
                f"Failed to launch updater:\n{e}" if self.lang == "EN" else f"Nie udało się uruchomić aktualizatora:\n{e}"
            )

    def add_new_book_form(self):
        form = tk.Toplevel(self)
        form.title("Add New Book" if self.lang == "EN" else "Dodaj nową książkę")
        form.geometry("460x390")
        form.grab_set()
        labels = {
            "EN": ["ID (0 may repeat)", "Title", "Author", "Year", "Genre", "Status", "Book Row"],
            "PL": ["ID (0 może się powtarzać)", "Tytuł", "Autor", "Rok", "Gatunek", "Status", "Regał książkowy"],
        }
        status_options = {
            "EN": ["borrowed", "available", "missing", "other"],
            "PL": ["wypożyczona", "dostępna", "brak", "inne"],
        }
        entries = {}
        for i, text in enumerate(labels[self.lang]):
            lbl = tk.Label(form, text=text, anchor="w")
            lbl.grid(row=i, column=0, padx=10, pady=8, sticky="w")
            if text == "Status" or text == "Status":
                cmb = ttk.Combobox(form, values=status_options[self.lang], state="readonly")
                cmb.grid(row=i, column=1, padx=10, pady=8, sticky="ew")
                cmb.current(1)  # Default to "available"
                entries[text] = cmb
            else:
                ent = tk.Entry(form)
                ent.grid(row=i, column=1, padx=10, pady=8, sticky="ew")
                entries[text] = ent
                ent.bind("<Return>", lambda e, i=i: self.entry_return_key(entries, labels[self.lang], i))
                ent.bind("<BackSpace>", lambda e, i=i: self.entry_backspace_key(e, entries, labels[self.lang], i))
                ent.bind("<Left>", lambda e, i=i: self.entry_arrow_key(e, entries, labels[self.lang], i, direction="left"))
                ent.bind("<Right>", lambda e, i=i: self.entry_arrow_key(e, entries, labels[self.lang], i, direction="right"))
                ent.bind("<Up>", lambda e, i=i: self.entry_arrow_key(e, entries, labels[self.lang], i, direction="up"))
                ent.bind("<Down>", lambda e, i=i: self.entry_arrow_key(e, entries, labels[self.lang], i, direction="down"))
        form.grid_columnconfigure(1, weight=1)
        form.after(100, lambda: entries[labels[self.lang][0]].focus_set())

        def submit():
            try:
                id_val = int(entries[labels[self.lang][0]].get())
                if id_val < 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror(
                    "Error" if self.lang == "EN" else "Błąd",
                    "ID must be a number >= 0" if self.lang == "EN" else "ID musi być liczbą >= 0",
                )
                return
            existing_ids = self.get_existing_ids()
            if id_val != 0 and id_val in existing_ids:
                messagebox.showerror(
                    "Error" if self.lang == "EN" else "Błąd",
                    "ID already exists or input 0 which may repeat" if self.lang == "EN" else "ID już istnieje lub wpisz 0, który może się powtarzać",
                )
                return
            title = entries[labels[self.lang][1]].get().strip()
            if not title:
                messagebox.showerror(
                    "Error" if self.lang == "EN" else "Błąd",
                    "Title is required" if self.lang == "EN" else "Tytuł jest wymagany",
                )
                return
            author = entries[labels[self.lang][2]].get().strip()
            year = entries[labels[self.lang][3]].get().strip()
            genre = entries[labels[self.lang][4]].get().strip()
            status = entries[labels[self.lang][5]].get().strip()
            book_row = entries[labels[self.lang][6]].get().strip()
            self.ws.append([id_val, title, author, year, genre, status, book_row])
            self.wb.save(self.file_name)
            messagebox.showinfo(
                "Success" if self.lang == "EN" else "Sukces",
                "Book added successfully." if self.lang == "EN" else "Książka dodana pomyślnie.",
            )
            form.destroy()

        submit_btn = tk.Button(
            form,
            text="Submit" if self.lang == "EN" else "Zatwierdź",
            bg=self.BTN_BG,
            fg=self.BTN_FG,
            activebackground=self.BTN_HOVER_BG,
            activeforeground=self.BTN_HOVER_FG,
            font=("Segoe UI", 12, "bold"),
            command=submit,
        )
        submit_btn.grid(row=len(labels[self.lang]), column=0, columnspan=2, pady=15, sticky="ew")
        submit_btn.bind("<Return>", lambda e: submit())

    def get_existing_ids(self):
        ids = []
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                ids.append(row[0])
        return ids

    def entry_return_key(self, entries, label_list, current_index):
        next_index = current_index + 1
        if next_index < len(label_list):
            entries[label_list[next_index]].focus_set()
        else:
            entries[label_list[current_index]].master.children["!button"].invoke()

    def entry_backspace_key(self, event, entries, label_list, current_index):
        entry = event.widget
        if entry.get() == "" and current_index > 0:
            entries[label_list[current_index - 1]].focus_set()
            return "break"

    def entry_arrow_key(self, event, entries, label_list, current_index, direction):
        if direction in ("left", "up") and current_index > 0:
            entries[label_list[current_index - 1]].focus_set()
        elif direction in ("right", "down") and current_index < len(label_list) - 1:
            entries[label_list[current_index + 1]].focus_set()
        return "break"

    def see_modify_books(self):
        top = tk.Toplevel(self)
        top.title("Books" if self.lang == "EN" else "Książki")
        top.geometry("900x520")
        search_frame = tk.Frame(top)
        search_frame.pack(fill="x", padx=5, pady=5)

        search_labels_en = [
            "Search by ID",
            "Search by Title",
            "Search by Author",
            "Search by Year",
            "Search by Genre",
            "Search by Status",
            "Search",
        ]
        search_labels_pl = [
            "Szukaj po ID",
            "Szukaj po tytule",
            "Szukaj po autorze",
            "Szukaj po roku",
            "Szukaj po gatunku",
            "Szukaj po statusie",
            "Szukaj",
        ]
        search_labels = search_labels_en if self.lang == "EN" else search_labels_pl
        self.search_vars = {key: tk.StringVar() for key in search_labels[:-1]}

        # FIX: Initialize sort directions for all columns including Status
        self.sort_directions = {key: True for key in search_labels[:-1]}

        entries_order = []

        for idx, label in enumerate(search_labels[:-1]):
            tk.Label(search_frame, text=label).grid(row=0, column=idx * 2, sticky="w")
            ent = tk.Entry(search_frame, textvariable=self.search_vars[label], width=15)
            ent.grid(row=0, column=idx * 2 + 1, padx=(0, 10), sticky="ew")
            entries_order.append(ent)
            ent.bind("<Return>", lambda e, i=idx: self.focus_or_search(i, entries_order, filter_tree))

            # Make columns resizable in search_frame
            search_frame.grid_columnconfigure(idx * 2, weight=0)
            search_frame.grid_columnconfigure(idx * 2 + 1, weight=1)

        btn_search = tk.Button(
            search_frame,
            text=search_labels[-1],
            bg=self.BTN_BG,
            fg=self.BTN_FG,
            activebackground=self.BTN_HOVER_BG,
            activeforeground=self.BTN_HOVER_FG,
            font=("Segoe UI", 11, "bold"),
            command=lambda: filter_tree(),
        )
        btn_search.grid(row=0, column=14, padx=(0, 10))
        search_frame.grid_columnconfigure(14, weight=0)

        cols = ("ID", "Title", "Author", "Year", "Genre", "Status")
        if self.lang == "EN":
            col_headings = ["ID", "Title", "Author", "Year", "Genre", "Status"]
        else:
            col_headings = ["ID", "Tytuł", "Autor", "Rok", "Gatunek", "Status"]
        tree = ttk.Treeview(top, columns=cols, show="headings")
        for col, heading in zip(cols, col_headings):
            tree.heading(col, text=heading, command=lambda c=col: self.sort_by_column(tree, c))
            tree.column(col, width=150 if col not in ("Year", "Status") else 80, stretch=True)
        tree.pack(expand=True, fill="both", padx=5, pady=5)

        top.grid_columnconfigure(0, weight=1)
        top.grid_rowconfigure(1, weight=1)  # Assuming tree packed in row 1 if used with grid, we pack so not mandatory to grid configure here

        # Allow buttons frame to expand
        btn_frame = tk.Frame(top)
        btn_frame.pack(fill="x", pady=5)
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(1, weight=1)
        btn_frame.grid_columnconfigure(2, weight=1)

        def load_tree_data(filter_criteria=None, sort_by=None, ascending=True):
            tree.delete(*tree.get_children())
            rows = list(self.ws.iter_rows(min_row=2, values_only=True))
            if filter_criteria:
                filtered = []
                for row in rows:
                    match = True
                    for key, val in filter_criteria.items():
                        if val:
                            idx = {"ID": 0, "Title": 1, "Author": 2, "Year": 3, "Genre": 4, "Status": 5}[key]
                            if key in ("ID", "Year"):
                                if str(val).casefold() != str(row[idx]).casefold():
                                    match = False
                                    break
                            else:
                                if val.casefold() not in (str(row[idx] or "")).casefold():
                                    match = False
                                    break
                    if match:
                        filtered.append(row)
                rows = filtered
            if sort_by:
                col_idx = {"ID": 0, "Title": 1, "Author": 2, "Year": 3, "Genre": 4, "Status": 5}[sort_by]

                def sort_key(r):
                    val = r[col_idx]
                    return str(val).casefold() if val is not None else ""

                rows.sort(key=sort_key, reverse=not ascending)
            for row in rows:
                row_vals = list(row)
                if len(row_vals) < 6:
                    row_vals += [""] * (6 - len(row_vals))
                tree.insert("", "end", values=row_vals)

        def filter_tree():
            crit = {}
            cleared_all = True
            for key in self.search_vars:
                val = self.search_vars[key].get().strip()
                if val != "":
                    cleared_all = False
                crit_key = key if self.lang == "EN" else key
                crit[crit_key] = val
            if cleared_all:
                load_tree_data()
            else:
                keys_map_pl_en = {
                    "Szukaj po ID": "ID",
                    "Szukaj po tytule": "Title",
                    "Szukaj po autorze": "Author",
                    "Szukaj po roku": "Year",
                    "Szukaj po gatunku": "Genre",
                    "Szukaj po statusie": "Status",
                    "Search by ID": "ID",
                    "Search by Title": "Title",
                    "Search by Author": "Author",
                    "Search by Year": "Year",
                    "Search by Genre": "Genre",
                    "Search by Status": "Status",
                }
                crit_en = {keys_map_pl_en.get(k, k): v for k, v in crit.items()}
                load_tree_data(crit_en)

        load_tree_data()

        def on_edit():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Warning" if self.lang == "EN" else "Uwaga",
                                       "Select a book to edit" if self.lang == "EN" else "Wybierz książkę do edycji")
                return
            item = tree.item(selected[0])
            values = item["values"]
            self.edit_book(values[0])
            top.destroy()

        def on_remove():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Warning" if self.lang == "EN" else "Uwaga",
                                       "Select a book to remove" if self.lang == "EN" else "Wybierz książkę do usunięcia")
                return
            item = tree.item(selected[0])
            values = item["values"]
            book_id = values[0]
            # Remove trailing book row, keep first 6 columns for removed list (ID,Title,Author,Year,Genre,Status)
            book_data = values[:6]
            confirmed = messagebox.askyesno(
                "Confirm removal" if self.lang == "EN" else "Potwierdź usunięcie",
                "Are you sure you want to remove the selected book?"
                if self.lang == "EN"
                else "Czy na pewno chcesz usunąć wybraną książkę?",
            )
            if not confirmed:
                return
            idx, row = find_book_by_id(self.ws, book_id)
            if idx is not None:
                self.ws.delete_rows(idx)
                self.wb.save(self.file_name)
            else:
                messagebox.showerror(
                    "Error" if self.lang == "EN" else "Błąd",
                    "Book not found." if self.lang == "EN" else "Nie znaleziono książki.",
                )
                return
            if os.path.exists(self.removed_file_name):
                wb_removed = load_workbook(self.removed_file_name)
                ws_removed = wb_removed.active
            else:
                wb_removed = Workbook()
                ws_removed = wb_removed.active
                ws_removed.title = "RemovedBooks"
                headers = ["ID", "Title", "Author", "Year", "Genre", "Status"]
                ws_removed.append(headers)
                for col_num in range(1, len(headers) + 1):
                    cell = ws_removed.cell(row=1, column=col_num)
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    ws_removed.column_dimensions[get_column_letter(col_num)].width = 15
            ws_removed.append(book_data)  # Save full 6 columns including Status
            wb_removed.save(self.removed_file_name)
            messagebox.showinfo(
                "Removed" if self.lang == "EN" else "Usunięto",
                "Book removed and saved in removed list."
                if self.lang == "EN"
                else "Książka usunięta i zapisana na liście usuniętych.",
            )
            tree.delete(selected[0])
            filter_tree()

        btn_edit = tk.Button(
            btn_frame,
            text="Edit Book" if self.lang == "EN" else "Edytuj książkę",
            bg=self.BTN_BG,
            fg=self.BTN_FG,
            activebackground=self.BTN_HOVER_BG,
            activeforeground=self.BTN_HOVER_FG,
            font=("Segoe UI", 11, "bold"),
            command=on_edit,
        )
        btn_edit.pack(side="left", padx=10)

        btn_remove = tk.Button(
            btn_frame,
            text="Remove Book" if self.lang == "EN" else "Usuń książkę",
            bg=self.BTN_BG,
            fg=self.BTN_FG,
            activebackground=self.BTN_HOVER_BG,
            activeforeground=self.BTN_HOVER_FG,
            font=("Segoe UI", 11, "bold"),
            command=on_remove,
        )
        btn_remove.pack(side="left", padx=10)

        btn_close = tk.Button(btn_frame, text="Close" if self.lang == "EN" else "Zamknij", command=top.destroy)
        btn_close.pack(side="bottom", fill="x", pady=(10, 0), padx=10)

    def focus_or_search(self, idx, entries, search_func):
        if idx + 1 < len(entries):
            entries[idx + 1].focus_set()
        else:
            search_func()

    def sort_by_column(self, tree, col):
        ascending = self.sort_directions[col]
        self.sort_directions[col] = not ascending
        crit = {}
        cleared_all = True
        for key in self.search_vars:
            val = self.search_vars[key].get().strip()
            keys_map_pl_en = {
                "Szukaj po ID": "ID",
                "Szukaj po tytule": "Title",
                "Szukaj po autorze": "Author",
                "Szukaj po roku": "Year",
                "Szukaj po gatunku": "Genre",
                "Szukaj po statusie": "Status",
                "Search by ID": "ID",
                "Search by Title": "Title",
                "Search by Author": "Author",
                "Search by Year": "Year",
                "Search by Genre": "Genre",
                "Search by Status": "Status",
            }
            crit[keys_map_pl_en.get(key, key)] = val
            if val != "":
                cleared_all = False
        filter_criteria = None if cleared_all else crit
        tree.delete(*tree.get_children())
        rows = list(self.ws.iter_rows(min_row=2, values_only=True))
        if filter_criteria:
            filtered = []
            for row in rows:
                match = True
                for key, val in filter_criteria.items():
                    if val:
                        col_i = {"ID": 0, "Title": 1, "Author": 2, "Year": 3, "Genre": 4, "Status": 5}[key]
                        if key in ("ID", "Year"):
                            val_cf = str(val).casefold()
                            cell_val_cf = str(row[col_i]).casefold() if row[col_i] is not None else ""
                            if val_cf != cell_val_cf:
                                match = False
                                break
                        else:
                            val_cf = str(val).casefold()
                            cell_val_cf = str(row[col_i]).casefold() if row[col_i] is not None else ""
                            if val_cf not in cell_val_cf:
                                match = False
                                break
                if match:
                    filtered.append(row)
            rows = filtered
        col_index = {"ID": 0, "Title": 1, "Author": 2, "Year": 3, "Genre": 4, "Status": 5}[col]
        def sort_key(x):
            val = x[col_index]
            return str(val).casefold() if val is not None else ""
        rows.sort(key=sort_key, reverse=not ascending)
        for row in rows:
            row_vals = list(row)
            if len(row_vals) < 6:
                row_vals += [""] * (6 - len(row_vals))
            tree.insert("", "end", values=row_vals)

    def edit_book(self, book_id):
        idx, row = find_book_by_id(self.ws, book_id)
        if row is None:
            messagebox.showerror(
                "Error" if self.lang == "EN" else "Błąd",
                "Book not found" if self.lang == "EN" else "Książka nie znaleziona",
            )
            return self.create_main_menu()
        current = {
            "ID": row[0].value,
            "Title": row[1].value,
            "Author": row[2].value,
            "Year": row[3].value,
            "Genre": row[4].value,
            "Status": (row[5].value if len(row) > 5 else "")
        }
        form = tk.Toplevel(self)
        form.title("Edit Book" if self.lang == "EN" else "Edytuj książkę")
        form.geometry("440x360")
        form.grab_set()
        labels = {
            "EN": ["ID (0 may repeat)", "Title", "Author", "Year", "Genre", "Status"],
            "PL": ["ID (0 może się powtarzać)", "Tytuł", "Autor", "Rok", "Gatunek", "Status"],
        }
        status_options = {
            "EN": ["borrowed", "available", "missing", "other"],
            "PL": ["wypożyczona", "dostępna", "brak", "inne"],
        }
        entries = {}
        values_current = [
            current["ID"],
            current["Title"],
            current["Author"],
            current["Year"],
            current["Genre"],
            current["Status"],
        ]
        for i, text in enumerate(labels[self.lang]):
            lbl = tk.Label(form, text=text, anchor="w")
            lbl.grid(row=i, column=0, padx=10, pady=8, sticky="w")
            if text == "Status" or text == "Status":
                cmb = ttk.Combobox(form, values=status_options[self.lang], state="readonly")
                cmb.grid(row=i, column=1, padx=10, pady=8, sticky="ew")
                try:
                    cmb.current(status_options[self.lang].index(str(values_current[i]).casefold()))
                except Exception:
                    cmb.current(1)  # default to "available"
                entries[text] = cmb
            else:
                ent = tk.Entry(form)
                ent.grid(row=i, column=1, padx=10, pady=8, sticky="ew")
                ent.insert(0, values_current[i])
                entries[text] = ent
                ent.bind("<Return>", lambda e, i=i: self.entry_return_key(entries, labels[self.lang], i))
                ent.bind("<BackSpace>", lambda e, i=i: self.entry_backspace_key(e, entries, labels[self.lang], i))
                ent.bind("<Left>", lambda e, i=i: self.entry_arrow_key(e, entries, labels[self.lang], i, direction="left"))
                ent.bind("<Right>", lambda e, i=i: self.entry_arrow_key(e, entries, labels[self.lang], i, direction="right"))
                ent.bind("<Up>", lambda e, i=i: self.entry_arrow_key(e, entries, labels[self.lang], i, direction="up"))
                ent.bind("<Down>", lambda e, i=i: self.entry_arrow_key(e, entries, labels[self.lang], i, direction="down"))
        form.grid_columnconfigure(1, weight=1)

        def submit():
            try:
                id_val = int(entries[labels[self.lang][0]].get())
                if id_val < 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror(
                    "Error" if self.lang == "EN" else "Błąd",
                    "ID must be a number >= 0" if self.lang == "EN" else "ID musi być liczbą >= 0",
                )
                return
            existing_ids = self.get_existing_ids()
            if id_val != 0 and id_val in existing_ids and id_val != current["ID"]:
                messagebox.showerror(
                    "Error" if self.lang == "EN" else "Błąd",
                    "ID already exists or input 0 which may repeat" if self.lang == "EN" else "ID już istnieje lub wpisz 0, który może się powtarzać",
                )
                return
            title = entries[labels[self.lang][1]].get().strip()
            if not title:
                messagebox.showerror(
                    "Error" if self.lang == "EN" else "Błąd",
                    "Title is required" if self.lang == "EN" else "Tytuł jest wymagany",
                )
                return
            author = entries[labels[self.lang][2]].get().strip()
            year = entries[labels[self.lang][3]].get().strip()
            genre = entries[labels[self.lang][4]].get().strip()
            status = entries[labels[self.lang][5]].get().strip()
            self.ws.cell(row=idx, column=1).value = id_val
            self.ws.cell(row=idx, column=2).value = title
            self.ws.cell(row=idx, column=3).value = author
            self.ws.cell(row=idx, column=4).value = year
            self.ws.cell(row=idx, column=5).value = genre
            self.ws.cell(row=idx, column=6).value = status
            self.wb.save(self.file_name)
            messagebox.showinfo(
                "Success" if self.lang == "EN" else "Sukces",
                "Book updated successfully." if self.lang == "EN" else "Książka zaktualizowana pomyślnie.",
            )
            form.destroy()

        submit_btn = tk.Button(
            form,
            text="Submit" if self.lang == "EN" else "Zatwierdź",
            bg=self.BTN_BG,
            fg=self.BTN_FG,
            activebackground=self.BTN_HOVER_BG,
            activeforeground=self.BTN_HOVER_FG,
            font=("Segoe UI", 12, "bold"),
            command=submit,
        )
        submit_btn.grid(row=len(labels[self.lang]), column=0, columnspan=2, pady=15, sticky="ew")
        submit_btn.bind("<Return>", lambda e: submit())

    def show_help(self):
        help_text_en = (
            "Bookworm Help:\n\n"
            "1. Add New Book: Add a new book with details.\n"
            "2. See/Modify Books: View and edit existing books.\n"
            "3. Help: Show this help message.\n"
            "4. Update: looks for the new program version. "
            "5. Settings: Change theme and default language, allows for custom themes.\n"
            "6. Exit: Close the program."
        )
        help_text_pl = (
            "Pomoc Bookworm:\n\n"
            "1. Dodaj książkę: Dodaj nową książkę z danymi.\n"
            "2. Obejrzyj/modyfikuj książki: Przeglądaj i edytuj książki.\n"
            "3. Pomoc: Wyświetl tę pomoc.\n"
            "4. Aktualizuj: wyszukuje nową wersję programu. \n"
            "5. Ustawienia: Zmień motyw i język domyślny, stwórz własny motyw.\n"
            "6. Zakończ: Zamknij program."
        )
        messagebox.showinfo("Help" if self.lang == "EN" else "Pomoc", help_text_en if self.lang == "EN" else help_text_pl)
        self.create_main_menu()

    def show_settings(self):
        win = tk.Toplevel(self)
        win.title("Settings" if self.lang == "EN" else "Ustawienia")
        win.geometry("400x300")
        win.grab_set()
        lbl_theme = tk.Label(win, text="Select Theme:" if self.lang == "EN" else "Wybierz motyw:")
        lbl_theme.grid(row=0, column=0, padx=10, pady=10, sticky="w")
        theme_var = tk.StringVar(value=self.settings.get("theme", "classic_blue"))
        theme_names = list(self.themes.keys())
        theme_cb = ttk.Combobox(win, values=theme_names, state="readonly", textvariable=theme_var)
        theme_cb.grid(row=0, column=1, padx=10, pady=10, sticky="ew")
        lbl_lang = tk.Label(win, text="Default Language:" if self.lang == "EN" else "Język domyślny:")
        lbl_lang.grid(row=1, column=0, padx=10, pady=10, sticky="w")
        lang_var = tk.StringVar(value=self.settings.get("default_language", "EN"))
        lang_cb = ttk.Combobox(win, values=["EN", "PL"], state="readonly", textvariable=lang_var)
        lang_cb.grid(row=1, column=1, padx=10, pady=10, sticky="ew")

        def save_settings():
            self.settings["theme"] = theme_var.get()
            self.settings["default_language"] = lang_var.get()
            self.save_settings(self.settings)
            self.load_custom_themes()
            self.set_theme(self.settings["theme"])
            messagebox.showinfo(
                "Settings Saved" if self.lang == "EN" else "Zapisano ustawienia",
                "Settings have been saved. Restart app to apply changes." if self.lang == "EN" else "Ustawienia zostały zapisane. Uruchom ponownie aplikację, aby zastosować zmiany.",
            )
            win.destroy()

        def run_theme_creator():
            themecreator_file, ver = find_latest_version_executable("themecreator")
            if not themecreator_file:
                messagebox.showinfo(
                    "Not Found" if self.lang == "EN" else "Nie znaleziono",
                    "Theme creator executable not found. Please install it." if self.lang == "EN" else "Nie znaleziono programu do tworzenia motywów. Proszę zainstalować.",
                )
                return
            try:
                subprocess.Popen([os.path.abspath(themecreator_file)], shell=True)
            except Exception as e:
                messagebox.showerror(
                    "Error" if self.lang == "EN" else "Błąd",
                    f"Failed to launch theme creator:\n{e}" if self.lang == "EN" else f"Nie udało się uruchomić programu do tworzenia motywów:\n{e}",
                )

        btn_save = tk.Button(
            win,
            text="Save" if self.lang == "EN" else "Zapisz",
            bg=self.BTN_BG,
            fg=self.BTN_FG,
            activebackground=self.BTN_HOVER_BG,
            activeforeground=self.BTN_HOVER_FG,
            font=("Segoe UI", 12, "bold"),
            command=save_settings,
        )
        btn_save.grid(row=3, column=0, columnspan=2, pady=30, sticky="ew")

        btn_theme_creator = tk.Button(
            win,
            text="Create New Theme" if self.lang == "EN" else "Nowy motyw",
            bg=self.BTN_BG,
            fg=self.BTN_FG,
            activebackground=self.BTN_HOVER_BG,
            activeforeground=self.BTN_HOVER_FG,
            font=("Segoe UI", 11, "bold"),
            command=run_theme_creator,
        )
        btn_theme_creator.grid(row=4, column=0, columnspan=2, pady=0, sticky="ew")

        win.grid_columnconfigure(1, weight=1)


if __name__ == "__main__":
    app = BookwormApp()
    app.mainloop()
