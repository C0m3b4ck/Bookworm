import os
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

FILE_NAME_EN = "books_en.xlsx"
FILE_NAME_PL = "books_pl.xlsx"
REMOVED_FILE_EN = "removed_en.xlsx"
REMOVED_FILE_PL = "removed_pl.xlsx"


class BookwormApp(tk.Tk):
    BTN_BG = "#add8e6"       # light blue
    BTN_FG = "#333333"       # dark grey
    BTN_HOVER_BG = "#6495ed" # darker blue (cornflower blue)
    BTN_HOVER_FG = "#f0f0f0" # lighter text

    def __init__(self):
        super().__init__()
        self.title("Bookworm")
        self.geometry("750x520")
        self.minsize(600, 450)

        self.lang = None
        self.file_name = None
        self.removed_file_name = None
        self.wb = None
        self.ws = None

        self.create_language_selection()

    def create_high_contrast_button(self, parent, text, command=None):
        btn = tk.Button(parent, text=text, bg=self.BTN_BG, fg=self.BTN_FG,
                        activebackground=self.BTN_HOVER_BG, activeforeground=self.BTN_HOVER_FG,
                        font=("Segoe UI", 12, "bold"), command=command, relief="raised", bd=3)
        btn.bind("<Enter>", lambda e: btn.config(bg=self.BTN_HOVER_BG, fg=self.BTN_HOVER_FG))
        btn.bind("<Leave>", lambda e: btn.config(bg=self.BTN_BG, fg=self.BTN_FG))
        return btn

    def create_language_selection(self):
        for widget in self.winfo_children():
            widget.destroy()

        tk.Label(self, text="Select language / wybierz język", font=("Arial", 18)).grid(row=0, column=0, columnspan=2, pady=20)

        btn_en = self.create_high_contrast_button(self, "English", lambda: self.start_app("EN"))
        btn_en.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)

        btn_pl = self.create_high_contrast_button(self, "Polski", lambda: self.start_app("PL"))
        btn_pl.grid(row=1, column=1, sticky="nsew", padx=10, pady=5)

        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)

    def start_app(self, lang):
        self.lang = lang
        self.file_name = FILE_NAME_EN if lang == "EN" else FILE_NAME_PL
        self.removed_file_name = REMOVED_FILE_EN if lang == "EN" else REMOVED_FILE_PL
        self.load_or_create_workbook()
        self.create_main_menu()

    def load_or_create_workbook(self):
        if not os.path.exists(self.file_name):
            if messagebox.askyesno("Create file" if self.lang=="EN" else "Utwórz plik",
                                   f"{'Excel file not found. Create it?' if self.lang=='EN' else 'Plik excela nie znaleziony. Utworzyć nowy?'}"):
                self.wb = Workbook()
                self.ws = self.wb.active
                self.ws.title = "Books"
                headers = ["ID", "Title", "Author", "Year", "Genre"]

                header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")

                self.ws.append(headers)
                for col_num in range(1, len(headers) + 1):
                    cell = self.ws.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    self.ws.column_dimensions[get_column_letter(col_num)].width = 15

                self.wb.save(self.file_name)
            else:
                messagebox.showinfo("Info", "Exiting." if self.lang=="EN" else "Kończenie programu.")
                self.quit()
                return

        self.wb = load_workbook(self.file_name)
        if "Books" not in self.wb.sheetnames:
            self.ws = self.wb.create_sheet("Books")
            headers = ["ID", "Title", "Author", "Year", "Genre"]
            header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            self.ws.append(headers)
            for col_num in range(1, len(headers)+1):
                cell = self.ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                self.ws.column_dimensions[get_column_letter(col_num)].width = 15
            self.wb.save(self.file_name)
        else:
            self.ws = self.wb["Books"]

        if not os.path.exists(self.removed_file_name):
            wb_removed = Workbook()
            ws_removed = wb_removed.active
            ws_removed.title = "RemovedBooks"
            headers = ["ID", "Title", "Author", "Year", "Genre"]
            ws_removed.append(headers)
            for col_num in range(1, len(headers) +1):
                cell = ws_removed.cell(row=1, column=col_num)
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                ws_removed.column_dimensions[get_column_letter(col_num)].width = 15
            wb_removed.save(self.removed_file_name)

    def create_main_menu(self):
        for widget in self.winfo_children():
            widget.destroy()

        credits_text = ("Made by C0m3b4ck under APL 2.0 license"
                        if self.lang == "EN"
                        else "Stworzone przez C0m3b4ck pod licencją APL 2.0")
        label_text = "Main Menu" if self.lang == "EN" else "Menu główne"
        tk.Label(self, text=label_text, font=("Arial", 20)).grid(row=0, column=0, columnspan=2, pady=10)
        tk.Label(self, text=credits_text, font=("Arial", 10, "italic")).grid(row=1, column=0, columnspan=2, pady=5)

        options = {
            "EN": ["Add New Book", "See/Modify Books", "Help", "Exit"],
            "PL": ["Dodaj książkę", "Obejrzyj/modyfikuj książki", "Pomoc", "Zakończ"]
        }

        for i, option in enumerate(options[self.lang], start=2):
            btn = self.create_high_contrast_button(self, option, command=lambda choice=i-1: self.handle_menu_choice(choice))
            btn.grid(row=i, column=0, columnspan=2, sticky="nsew", padx=50, pady=7)

        for i in range(2, 6):
            self.grid_rowconfigure(i, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

    def handle_menu_choice(self, choice):
        if choice == 1:
            self.add_new_book_form()
        elif choice == 2:
            self.see_modify_books()
        elif choice == 3:
            self.show_help()
        elif choice == 4:
            self.quit()

    def get_existing_ids(self):
        return [row[0] for row in self.ws.iter_rows(min_row=2, values_only=True) if row[0] is not None]

    def add_new_book_form(self):
        form = tk.Toplevel(self)
        form.title("Add New Book" if self.lang == "EN" else "Dodaj nową książkę")
        form.geometry("400x320")
        form.grab_set()

        labels = {
            "EN": ["ID (0 may repeat)", "Title", "Author", "Year", "Genre"],
            "PL": ["ID (0 może się powtarzać)", "Tytuł", "Autor", "Rok", "Gatunek"]
        }
        entries = {}

        for i, text in enumerate(labels[self.lang]):
            lbl = tk.Label(form, text=text, anchor="w")
            lbl.grid(row=i, column=0, padx=10, pady=8, sticky="w")
            ent = tk.Entry(form)
            ent.grid(row=i, column=1, padx=10, pady=8, sticky="ew")
            entries[text] = ent

        form.grid_columnconfigure(1, weight=1)

        def submit():
            try:
                id_val = int(entries[labels[self.lang][0]].get())
                if id_val < 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Error" if self.lang == "EN" else "Błąd",
                                     "ID must be a number >= 0" if self.lang == "EN" else "ID musi być liczbą >= 0")
                return
            existing_ids = self.get_existing_ids()
            if id_val != 0 and id_val in existing_ids:
                messagebox.showerror("Error" if self.lang == "EN" else "Błąd",
                                     "ID already exists or input 0 which may repeat" if self.lang == "EN" else "ID już istnieje lub wpisz 0, który może się powtarzać")
                return
            title = entries[labels[self.lang][1]].get().strip()
            if not title:
                messagebox.showerror("Error" if self.lang == "EN" else "Błąd",
                                     "Title is required" if self.lang == "EN" else "Tytuł jest wymagany")
                return
            author = entries[labels[self.lang][2]].get().strip()
            year = entries[labels[self.lang][3]].get().strip()
            genre = entries[labels[self.lang][4]].get().strip()

            self.ws.append([id_val, title, author, year, genre])
            self.wb.save(self.file_name)
            messagebox.showinfo("Success" if self.lang == "EN" else "Sukces",
                                "Book added successfully." if self.lang == "EN" else "Książka dodana pomyślnie.")
            form.destroy()

        submit_btn = tk.Button(form, text="Submit" if self.lang == "EN" else "Zatwierdź",
                               bg=self.BTN_BG, fg=self.BTN_FG,
                               activebackground=self.BTN_HOVER_BG, activeforeground=self.BTN_HOVER_FG,
                               font=("Segoe UI", 12, "bold"),
                               command=submit)
        submit_btn.grid(row=len(labels[self.lang]), column=0, columnspan=2, pady=15, sticky="ew")

    def see_modify_books(self):
        top = tk.Toplevel(self)
        top.title("Books" if self.lang == "EN" else "Książki")
        top.geometry("790x520")

        search_frame = tk.Frame(top)
        search_frame.pack(fill='x', padx=5, pady=5)

        search_labels = {
            "EN": ["Search by ID", "Search by Title", "Search by Author", "Search by Year", "Search by Genre", "Search"],
            "PL": ["Szukaj po ID", "Szukaj po tytule", "Szukaj po autorze", "Szukaj po roku", "Szukaj po gatunku", "Szukaj"]
        }

        search_vars = {key: tk.StringVar() for key in search_labels[self.lang][:-1]}

        for idx, label in enumerate(search_labels[self.lang][:-1]):
            tk.Label(search_frame, text=label).grid(row=0, column=idx*2, sticky="w")
            tk.Entry(search_frame, textvariable=search_vars[search_labels["EN" if self.lang=="EN" else "PL"][idx]], width=15).grid(row=0, column=idx*2+1, padx=(0,10), sticky="w")

        btn_search = tk.Button(search_frame, text=search_labels[self.lang][-1], bg=self.BTN_BG, fg=self.BTN_FG,
                               activebackground=self.BTN_HOVER_BG, activeforeground=self.BTN_HOVER_FG,
                               font=("Segoe UI", 11, "bold"),
                               command=lambda: filter_tree())
        btn_search.grid(row=0, column=10, padx=(0,10))

        cols = ("ID", "Title", "Author", "Year", "Genre")
        tree = ttk.Treeview(top, columns=cols, show='headings')
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=150 if col != "Year" else 80, stretch=True)
        tree.pack(expand=True, fill='both', padx=5, pady=5)

        def load_tree_data(filter_criteria=None):
            tree.delete(*tree.get_children())
            for row in self.ws.iter_rows(min_row=2, values_only=True):
                if filter_criteria:
                    match = True
                    for key, val in filter_criteria.items():
                        if val:
                            idx = {"ID":0, "Title":1, "Author":2, "Year":3, "Genre":4}[key]
                            if val.lower() not in str(row[idx]).lower():
                                match = False
                                break
                    if match:
                        tree.insert("", "end", values=row)
                else:
                    tree.insert("", "end", values=row)

        def filter_tree():
            crit = {}
            if self.lang == "EN":
                crit = {
                    "ID": search_vars["Search by ID"].get().strip(),
                    "Title": search_vars["Search by Title"].get().strip(),
                    "Author": search_vars["Search by Author"].get().strip(),
                    "Year": search_vars["Search by Year"].get().strip(),
                    "Genre": search_vars["Search by Genre"].get().strip()
                }
            else:
                crit = {
                    "ID": search_vars["Szukaj po ID"].get().strip(),
                    "Title": search_vars["Szukaj po tytule"].get().strip(),
                    "Author": search_vars["Szukaj po autorze"].get().strip(),
                    "Year": search_vars["Szukaj po roku"].get().strip(),
                    "Genre": search_vars["Szukaj po gatunku"].get().strip()
                }
            load_tree_data(crit)

        load_tree_data()

        btn_frame = tk.Frame(top)
        btn_frame.pack(fill='x', pady=5)

        def on_edit():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Warning" if self.lang == "EN" else "Uwaga",
                                       "Select a book to edit" if self.lang == "EN" else "Wybierz książkę do edycji")
                return
            item = tree.item(selected[0])
            values = item['values']
            self.edit_book(values[0])
            top.destroy()

        def on_remove():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Warning" if self.lang == "EN" else "Uwaga",
                                       "Select a book to remove" if self.lang == "EN" else "Wybierz książkę do usunięcia")
                return
            item = tree.item(selected[0])
            values = item['values']
            book_id = values[0]
            book_data = values

            confirmed = messagebox.askyesno("Confirm removal" if self.lang == "EN" else "Potwierdź usunięcie",
                                            "Are you sure you want to remove the selected book?" if self.lang == "EN" else "Czy na pewno chcesz usunąć wybraną książkę?")
            if not confirmed:
                return

            idx, row = find_book_by_id(self.ws, book_id)
            if idx is not None:
                self.ws.delete_rows(idx)
                self.wb.save(self.file_name)
            else:
                messagebox.showerror("Error" if self.lang == "EN" else "Błąd",
                                     "Book not found." if self.lang == "EN" else "Nie znaleziono książki.")
                return

            # Save to removed file
            if os.path.exists(self.removed_file_name):
                wb_removed = load_workbook(self.removed_file_name)
                ws_removed = wb_removed.active
            else:
                wb_removed = Workbook()
                ws_removed = wb_removed.active
                ws_removed.title = "RemovedBooks"
                headers = ["ID", "Title", "Author", "Year", "Genre"]
                ws_removed.append(headers)
                for col_num in range(1, len(headers) +1):
                    cell = ws_removed.cell(row=1, column=col_num)
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    ws_removed.column_dimensions[get_column_letter(col_num)].width = 15

            ws_removed.append(book_data)
            wb_removed.save(self.removed_file_name)

            messagebox.showinfo("Removed" if self.lang == "EN" else "Usunięto",
                                "Book removed and saved in removed list." if self.lang == "EN" else "Książka usunięta i zapisana na liście usuniętych.")
            tree.delete(selected[0])

        btn_edit = tk.Button(btn_frame, text="Edit Book" if self.lang == "EN" else "Edytuj książkę",
                             bg=self.BTN_BG, fg=self.BTN_FG,
                             activebackground=self.BTN_HOVER_BG, activeforeground=self.BTN_HOVER_FG,
                             font=("Segoe UI", 11, "bold"), command=on_edit)
        btn_edit.pack(side='left', padx=10)

        btn_remove = tk.Button(btn_frame, text="Remove Book" if self.lang == "EN" else "Usuń książkę",
                               bg=self.BTN_BG, fg=self.BTN_FG,
                               activebackground=self.BTN_HOVER_BG, activeforeground=self.BTN_HOVER_FG,
                               font=("Segoe UI", 11, "bold"), command=on_remove)
        btn_remove.pack(side='left', padx=10)

        btn_close = tk.Button(btn_frame, text="Close" if self.lang == "EN" else "Zamknij",
                              command=top.destroy)
        btn_close.pack(side='right', padx=10)

    def edit_book(self, book_id):
        idx, row = find_book_by_id(self.ws, book_id)
        if row is None:
            messagebox.showerror("Error" if self.lang == "EN" else "Błąd",
                                 "Book not found" if self.lang == "EN" else "Książka nie znaleziona")
            return self.create_main_menu()

        current = {
            "ID": row[0].value,
            "Title": row[1].value,
            "Author": row[2].value,
            "Year": row[3].value,
            "Genre": row[4].value
        }

        form = tk.Toplevel(self)
        form.title("Edit Book" if self.lang == "EN" else "Edytuj książkę")
        form.geometry("400x320")
        form.grab_set()

        labels = {
            "EN": ["ID (0 may repeat)", "Title", "Author", "Year", "Genre"],
            "PL": ["ID (0 może się powtarzać)", "Tytuł", "Autor", "Rok", "Gatunek"]
        }
        entries = {}

        values_current = [current["ID"], current["Title"], current["Author"], current["Year"], current["Genre"]]

        for i, text in enumerate(labels[self.lang]):
            lbl = tk.Label(form, text=text, anchor="w")
            lbl.grid(row=i, column=0, padx=10, pady=8, sticky="w")
            ent = tk.Entry(form)
            ent.grid(row=i, column=1, padx=10, pady=8, sticky="ew")
            ent.insert(0, values_current[i])
            entries[text] = ent

        form.grid_columnconfigure(1, weight=1)

        def submit():
            try:
                id_val = int(entries[labels[self.lang][0]].get())
                if id_val < 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Error" if self.lang == "EN" else "Błąd",
                                     "ID must be a number >= 0" if self.lang == "EN" else "ID musi być liczbą >= 0")
                return
            existing_ids = self.get_existing_ids()
            if id_val != 0 and id_val in existing_ids and id_val != current["ID"]:
                messagebox.showerror("Error" if self.lang == "EN" else "Błąd",
                                     "ID already exists or input 0 which may repeat" if self.lang == "EN" else "ID już istnieje lub wpisz 0, który może się powtarzać")
                return

            title = entries[labels[self.lang][1]].get().strip()
            if not title:
                messagebox.showerror("Error" if self.lang == "EN" else "Błąd",
                                     "Title is required" if self.lang == "EN" else "Tytuł jest wymagany")
                return

            author = entries[labels[self.lang][2]].get().strip()
            year = entries[labels[self.lang][3]].get().strip()
            genre = entries[labels[self.lang][4]].get().strip()

            # Update
            self.ws.cell(row=idx, column=1).value = id_val
            self.ws.cell(row=idx, column=2).value = title
            self.ws.cell(row=idx, column=3).value = author
            self.ws.cell(row=idx, column=4).value = year
            self.ws.cell(row=idx, column=5).value = genre

            self.wb.save(self.file_name)
            messagebox.showinfo("Success" if self.lang == "EN" else "Sukces",
                                "Book updated successfully." if self.lang == "EN" else "Książka zaktualizowana pomyślnie.")
            form.destroy()

        submit_btn = tk.Button(form, text="Submit" if self.lang == "EN" else "Zatwierdź",
                               bg=self.BTN_BG, fg=self.BTN_FG,
                               activebackground=self.BTN_HOVER_BG, activeforeground=self.BTN_HOVER_FG,
                               font=("Segoe UI", 12, "bold"), command=submit)
        submit_btn.grid(row=len(labels[self.lang]), column=0, columnspan=2, pady=15, sticky="ew")

    def show_help(self):
        help_text_en = ("Bookworm Help:\n\n"
                        "1. Add New Book: Add a new book with details.\n"
                        "2. See/Modify Books: View and edit existing books.\n"
                        "3. Help: Show this help message.\n"
                        "4. Exit: Close the program.")
        help_text_pl = ("Pomoc Bookworm:\n\n"
                        "1. Dodaj książkę: Dodaj nową książkę z danymi.\n"
                        "2. Obejrzyj/modyfikuj książki: Przeglądaj i edytuj książki.\n"
                        "3. Pomoc: Wyświetl tę pomoc.\n"
                        "4. Zakończ: Zamknij program.")
        messagebox.showinfo("Help" if self.lang == "EN" else "Pomoc",
                            help_text_en if self.lang == "EN" else help_text_pl)
        self.create_main_menu()

def find_book_by_id(ws, id_):
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        cell_id = row[0].value
        if cell_id == id_:
            return idx, row
    return None, None

if __name__ == "__main__":
    app = BookwormApp()
    app.mainloop()
